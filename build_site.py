#!/usr/bin/env python3
"""
UMBRA LEACHIES — site generator
--------------------------------
Reads leachianus_collection_EN_v2.xlsx and writes the public pages:

  index.html        home
  records.html      collection table
  lch-01.html …     one page per animal (leachianus + hognose)
  available.html    planned pairings / waitlist

Run from the folder that holds the workbook:

    python build_site.py                 # writes into ./site/
    python build_site.py --out umbra     # or straight into your Pages folder

Requires:  pip install openpyxl

Everything editable lives in the CONFIG block below. The workbook is the only
data source — change a weight or note in Excel, re-run, done.
"""
import argparse, datetime, html, os, re, sys
from openpyxl import load_workbook

# ============================== CONFIG =====================================
WORKBOOK   = "leachianus_collection_EN_v2.xlsx"
SITE_URL   = "https://jeremyngyintai-hub.github.io/umbra/"
ICON_FILE  = "umbra-icon.svg"          # copied next to the pages
PHOTO_DIR  = "photos"                  # photos/lch-01.jpg — optional, page hides the slot if missing
OG_IMAGE   = "photos/umbra-og.jpg"     # share preview image (make one ~1200×630)

CONTACT = {
    "whatsapp": "",                    # e.g. "https://wa.me/852XXXXXXXX"
    "instagram": "",                   # e.g. "https://instagram.com/umbraleachies"
    "email": "",                       # e.g. "mailto:umbra@example.com"
}

# Quarantine end dates by animal ID (remove an entry when quarantine ends)
QUARANTINE = {
    "LCH-01": "2026-09-19", "LCH-02": "2026-09-19",
    "LCH-05": "2026-09-25", "LCH-07": "2026-09-25", "LCH-09": "2026-09-25",
}

# Provenance claims and their status. Shown as chips on each record.
#   ("verified"|"unverified", EN text, ZH text)
CLAIMS = {
    "LCH-01": [("verified",   "Visual melanistic — MorphMarket 2127896, Ridiculous Rhacs",
                              "顯性黑化 — MorphMarket 2127896，Ridiculous Rhacs")],
    "LCH-02": [("unverified", "Possible het melanistic — Koghis-derived lines; parents not yet documented",
                              "可能帶黑化 — Koghis 血系；父母未有紀錄")],
    "LCH-03": [("verified",   "LL Pure Caanawa — Leachie Kingdom", "LL 純 Caanawa — Leachie Kingdom")],
    "LCH-04": [("verified",   "LL Pure Caanawa — Leachie Kingdom", "LL 純 Caanawa — Leachie Kingdom")],
    "LCH-05": [("verified",   "Nuu Ana — tub label, Leachie Kingdom", "Nuu Ana — 原飼養標籤，Leachie Kingdom"),
               ("unverified", "\"Possible het mel\" — seller's note only; conflicts with a pure island locality",
                              "「可能帶黑化」— 只係賣家備註；同純島型 locality 有矛盾")],
    "LCH-07": [("verified",   "Nuu Ami × Nuu Ami — tub label, Leachie Kingdom", "Nuu Ami × Nuu Ami — 原飼養標籤，Leachie Kingdom")],
    "LCH-09": [("verified",   "Brosse — tub label with laid / hatch dates", "Brosse — 原飼養標籤連產卵 / 孵化日")],
    "LCH-10": [("unverified", "\"Type D\" — no GT Type D exists; may be Isle D (Brosse) cross",
                              "「Type D」— 冇 GT Type D 呢個分類；可能係 Isle D（Brosse）雜交")],
}

# Named breeders worth showing (private sellers are not shown)
BREEDERS = {"Ridiculous": "Ridiculous Rhacs (US)", "Leachie Kingdom": "Leachie Kingdom"}

# Expected pairings for the Available page: (code, sire, dam, expected, EN, ZH)
# Leave empty to pull straight from the Breeding sheet.
AVAILABLE_OVERRIDE = []
# ===========================================================================

TODAY = datetime.date.today().isoformat()

def esc(x):  return html.escape(str(x)) if x not in (None, "") else "—"
def dt(x):   return x.strftime("%Y-%m-%d") if hasattr(x, "strftime") else "—"
def nm(s):   return re.sub(r"\s+(1\.0|0\.1)\s*$", "", str(s or "")).strip()
def sym(sex): return "♂" if sex == "1.0" else "♀"

# ------------------------------------------------------------------ data
def load(path):
    wb = load_workbook(path, data_only=True)
    c, w, br = wb["Collection"], wb["Weight Log"], wb["Breeding"]
    animals = {}
    for r in range(5, c.max_row + 1):
        i = c.cell(r, 1).value
        if not i or not str(i).startswith(("LCH", "HOG")): continue
        src = str(c.cell(r, 13).value or "")
        og = next((v for k, v in BREEDERS.items() if k in src), "")
        animals[i] = dict(name=c.cell(r, 2).value, lin=c.cell(r, 3).value, sex=c.cell(r, 4).value,
                          laid=c.cell(r, 5).value, hatch=c.cell(r, 6).value, age=c.cell(r, 7).value,
                          wt=c.cell(r, 8).value, dry=c.cell(r, 9).value, prep=c.cell(r, 10).value,
                          freq=c.cell(r, 11).value, acq=c.cell(r, 12).value, og=og,
                          reg=c.cell(r, 14).value, val=c.cell(r, 16).value, notes=c.cell(r, 17).value or "")
    weights = {}
    for r in range(5, w.max_row + 1):
        i, d, v = w.cell(r, 2).value, w.cell(r, 1).value, w.cell(r, 4).value
        if i and v is not None and hasattr(d, "strftime"):
            weights.setdefault(i, []).append((d, float(v)))
    for k in weights: weights[k].sort()
    pairs = []
    for r in range(5, 40):
        code, sire, dam = br.cell(r, 1).value, br.cell(r, 2).value, br.cell(r, 3).value
        if isinstance(code, str) and re.match(r"P-\d", code) and isinstance(sire, str):
            pairs.append(dict(code=code, sire=sire, dam=dam or "", goal=br.cell(r, 4).value or "",
                              sire_ready=br.cell(r, 5).value, dam_ready=br.cell(r, 6).value,
                              status=br.cell(r, 7).value or ""))
    return animals, weights, pairs

# ------------------------------------------------------------------ html
CSS = r"""
:root{--ink:#080D0E;--deep:#050809;--jade:#2EE6A8;--pale:#6FF3C6;--slate:#5A736F;--bone:#C4D6D2;
--ochre:#C9A227;--line:#183430;--display:"Cormorant Garamond",Georgia,serif;
--data:"IBM Plex Mono",ui-monospace,monospace;--body:"Inter",system-ui,sans-serif;
--cal:"Ma Shan Zheng","Noto Serif CJK HK",serif;--zh:"Noto Sans HK","PingFang HK","Microsoft JhengHei",sans-serif}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--ink);color:var(--bone);font-family:var(--body);font-size:15px;line-height:1.62;padding:0 20px 70px}
body.zh{font-family:var(--zh)}
.wrap{max-width:740px;margin:0 auto}
a{color:var(--jade);text-decoration:none}
.top{display:flex;align-items:center;gap:12px;padding:20px 0;border-bottom:1px solid var(--line);flex-wrap:wrap}
.top svg{width:28px;height:28px}
.top .b{font-family:var(--display);font-size:13px;letter-spacing:.3em;color:var(--jade)}
.top .z{font-family:var(--cal);font-size:17px;color:var(--pale)}
nav{margin-left:auto;display:flex;gap:18px;align-items:center}
nav a{font-family:var(--data);font-size:10px;letter-spacing:.16em;text-transform:uppercase;color:var(--slate)}
nav a.on{color:var(--pale)}
.lang{font-family:var(--data);font-size:10px;letter-spacing:.1em;border:1px solid var(--line);padding:4px 9px;color:var(--pale);cursor:pointer;background:var(--deep)}
[lang-en],[lang-zh]{display:none}
body.en [lang-en]{display:revert} body.zh [lang-zh]{display:revert}
body.en span[lang-en],body.en a[lang-en],body.en em[lang-en]{display:inline}
body.zh span[lang-zh],body.zh a[lang-zh],body.zh em[lang-zh]{display:inline}
header{padding:46px 0 34px;border-bottom:1px solid var(--line)}
.id{font-family:var(--data);font-size:11px;letter-spacing:.3em;color:var(--jade)}
h1{font-family:var(--display);font-weight:400;font-size:clamp(30px,7vw,46px);line-height:1.12;margin:10px 0 6px}
body.zh h1{font-family:var(--zh);font-weight:500;font-size:clamp(26px,6vw,38px)}
h1 .sx{font-family:var(--data);font-size:.5em;color:var(--jade);vertical-align:.18em}
.sub{font-family:var(--data);font-size:11px;letter-spacing:.12em;color:var(--slate);text-transform:uppercase}
.chips{margin-top:18px;display:flex;flex-wrap:wrap;gap:8px}
.chip{font-family:var(--data);font-size:10px;letter-spacing:.12em;text-transform:uppercase;border:1px solid var(--line);padding:5px 10px;color:var(--pale);background:var(--deep)}
.chip.q{border-color:#4A3C10;color:var(--ochre)}
.chip.v{border-color:#1E6B4E;color:var(--jade)} .chip.v::before{content:"✓ "}
.chip.u{border-color:#4A3C10;color:var(--ochre)} .chip.u::before{content:"? "}
section{padding:34px 0;border-bottom:1px solid var(--line)}
.eyebrow{font-family:var(--data);font-size:9.5px;letter-spacing:.3em;text-transform:uppercase;color:var(--jade);margin-bottom:16px}
h2{font-family:var(--display);font-weight:400;font-size:28px;color:var(--bone);margin-bottom:14px;line-height:1.2}
body.zh h2{font-family:var(--zh);font-weight:500;font-size:24px}
p{max-width:62ch} p+p{margin-top:12px} .muted{color:var(--slate)}
table{width:100%;border-collapse:collapse;font-size:14px}
td,th{padding:8px 0;border-bottom:1px solid #0F211E;text-align:left;vertical-align:top}
th{font-family:var(--data);font-size:10px;letter-spacing:.14em;text-transform:uppercase;color:var(--slate);font-weight:400;width:40%}
body.zh th{font-family:var(--zh);letter-spacing:.05em;font-size:12px}
td.n{font-family:var(--data);text-align:right;white-space:nowrap}
.spark{width:100%;height:150px;margin:8px 0 18px}
.notes{color:var(--slate);font-size:14px;max-width:66ch}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:6px}
.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.card{display:block;background:var(--deep);border:1px solid var(--line);padding:16px;color:var(--bone)}
.card .k{font-family:var(--data);font-size:9.5px;letter-spacing:.16em;text-transform:uppercase;color:var(--slate)}
.card .v{font-family:var(--display);font-size:26px;color:var(--pale);margin-top:4px}
.card .t{font-size:13px;color:var(--bone);margin-top:6px;line-height:1.5}
.photo{margin:0 0 6px;background:var(--deep);border:1px solid var(--line);aspect-ratio:3/2;overflow:hidden}
.photo img{width:100%;height:100%;object-fit:cover;display:block}
.cta{display:inline-block;margin:6px 10px 0 0;font-family:var(--data);font-size:11px;letter-spacing:.14em;text-transform:uppercase;border:1px solid var(--jade);padding:10px 16px;color:var(--jade)}
.hero{padding:72px 0 54px;text-align:center;border-bottom:1px solid var(--line)}
.hero svg{width:110px;height:110px;margin:0 auto 30px;display:block}
.hero .w{font-family:var(--display);font-size:clamp(40px,12vw,76px);letter-spacing:.3em;text-indent:.3em;color:var(--jade);line-height:.95}
.hero .w2{font-family:var(--display);font-size:clamp(18px,5vw,30px);letter-spacing:.42em;text-indent:.42em;color:var(--jade);opacity:.9;margin-top:8px}
.hero .zhw{font-family:var(--cal);font-size:clamp(42px,12vw,68px);color:var(--pale);margin-top:22px;letter-spacing:.14em;text-indent:.14em}
.hero .tag{font-family:var(--data);font-size:10px;letter-spacing:.24em;text-transform:uppercase;color:var(--slate);margin-top:22px}
footer{padding:26px 0;font-family:var(--data);font-size:10px;letter-spacing:.14em;color:var(--slate);text-transform:uppercase}
@media(max-width:600px){.grid2,.grid3{grid-template-columns:1fr}th{width:48%}nav{gap:12px}}
"""

JS = r"""
(function(){var k='umbra-lang',l=localStorage.getItem(k)||((navigator.language||'').startsWith('zh')?'zh':'en');
function set(x){document.body.className=x;localStorage.setItem(k,x);document.documentElement.lang=(x==='zh'?'zh-HK':'en');
var b=document.getElementById('langbtn');if(b)b.textContent=(x==='zh'?'EN':'中文');}
set(l);var b=document.getElementById('langbtn');if(b)b.onclick=function(){set(document.body.className==='zh'?'en':'zh');};})();
"""

def bi(en, zh, tag="span"):
    return f'<{tag} lang-en>{en}</{tag}><{tag} lang-zh>{zh}</{tag}>'

def head(title, page, desc="", og=None, icon=""):
    og = og or OG_IMAGE
    nav = "".join(
        f'<a href="./{f}" class="{"on" if page==f else ""}">{bi(en,zh)}</a>'
        for f, en, zh in [("index.html","Home","主頁"),("records.html","Records","紀錄"),
                          ("available.html","Available","預訂"),("brand.html","Brand","品牌")])
    return (f'<!DOCTYPE html><html lang="en"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width, initial-scale=1"><title>{esc(title)}</title>'
            f'<meta name="description" content="{esc(desc)}">'
            f'<meta property="og:title" content="{esc(title)}"><meta property="og:description" content="{esc(desc)}">'
            f'<meta property="og:image" content="{SITE_URL}{og}"><meta property="og:type" content="website">'
            f'<link rel="icon" type="image/svg+xml" href="{ICON_FILE}">'
            '<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
            '<link href="https://fonts.googleapis.com/css2?family=Ma+Shan+Zheng&family=Cormorant+Garamond:wght@300;400;500&family=IBM+Plex+Mono:wght@400;500&family=Inter:wght@400;500&family=Noto+Sans+HK:wght@400;500&display=swap" rel="stylesheet">'
            f'<style>{CSS}</style></head><body class="en"><div class="wrap">'
            f'<div class="top">{icon}<span class="b">UMBRA LEACHIES</span><span class="z">玄影</span>'
            f'<nav>{nav}<button class="lang" id="langbtn">中文</button></nav></div>')

def foot():
    return f'<footer>UMBRA LEACHIES · 玄影 · Hong Kong · {bi("updated","更新")} {TODAY}</footer></div><script>{JS}</script></body></html>'

def contact_block():
    links = [(k, v) for k, v in CONTACT.items() if v]
    if not links:
        return f'<p class="muted">{bi("Contact details to be added.","聯絡方式稍後補上。")}</p>'
    lab = {"whatsapp": "WhatsApp", "instagram": "Instagram", "email": "Email"}
    return "".join(f'<a class="cta" href="{v}">{lab[k]}</a>' for k, v in links)

def spark(pts):
    if len(pts) < 2: return ""
    xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
    x0, x1, y0, y1 = min(xs), max(xs), min(ys), max(ys)
    span = (x1 - x0).days or 1; yr = (y1 - y0) or 1; W, H = 560, 150
    pl = [(40 + (p[0]-x0).days/span*(W-60), H-20-(p[1]-y0)/yr*(H-45)) for p in pts]
    line = " ".join(f"{'M' if k==0 else 'L'}{x:.1f} {y:.1f}" for k, (x, y) in enumerate(pl))
    dots = "".join(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="#2EE6A8"/>' for x, y in pl)
    return (f'<svg viewBox="0 0 {W} {H}" class="spark" preserveAspectRatio="none">'
            f'<path d="{line}" fill="none" stroke="#2EE6A8" stroke-width="2"/>{dots}'
            f'<text x="4" y="16" fill="#5A736F" font-size="11">{y1:g} g</text>'
            f'<text x="4" y="{H-16}" fill="#5A736F" font-size="11">{y0:g} g</text></svg>')

# ------------------------------------------------------------------ pages
def animal_page(i, a, weights, pairs, icon):
    pts = weights.get(i, [])
    chips = f'<span class="chip">{esc(a["sex"])}</span>'
    if a["og"]: chips += f'<span class="chip">{esc(a["og"])}</span>'
    if a["reg"] and str(a["reg"]) != "N/A": chips += f'<span class="chip">MorphMarket {esc(a["reg"])}</span>'
    if i in QUARANTINE: chips += f'<span class="chip q">{bi("Quarantine to","隔離至")} {QUARANTINE[i]}</span>'
    claims = "".join(f'<span class="chip {"v" if s=="verified" else "u"}">{bi(en,zh)}</span>' for s, en, zh in CLAIMS.get(i, []))
    rows = [(("Lineage / locality","血統 / 產地"), esc(a["lin"])), (("Sex","性別"), esc(a["sex"])),
            (("Laid","產卵"), dt(a["laid"])), (("Hatched","孵化"), dt(a["hatch"])), (("Age","年齡"), esc(a["age"])),
            (("Acquired","購入"), dt(a["acq"])), (("Origin","來源"), esc(a["og"]) if a["og"] else "—")]
    tbl = "".join(f'<tr><th>{bi(*k)}</th><td>{v}</td></tr>' for k, v in rows)
    latest = pts[-1] if pts else None
    delta = f"{latest[1]-pts[-2][1]:+g} g" if len(pts) > 1 else "—"
    cards = (f'<div class="grid2"><div class="card"><div class="k">{bi("Current weight","現時體重")}</div><div class="v">{latest[1]:g} g</div></div>'
             f'<div class="card"><div class="k">{bi("Change at last weigh-in","對上一次變化")}</div><div class="v">{delta}</div></div></div>') if latest else ""
    wl = "".join(f'<tr><th>{dt(d)}</th><td class="n">{v:g} g</td></tr>' for d, v in reversed(pts))
    if isinstance(a["dry"], (int, float)):
        feed = (f'<tr><th>{bi("Dry powder","乾粉")}</th><td class="n">{a["dry"]:g} g</td></tr>'
                f'<tr><th>{bi("Mixed","果泥")}</th><td class="n">{a["prep"]:g} g</td></tr>'
                f'<tr><th>{bi("Frequency","頻率")}</th><td class="n">{esc(a["freq"])}</td></tr>')
    else:
        feed = f'<tr><th>{bi("Diet","飲食")}</th><td>{bi("See snake feeding record","見蛇類餵食紀錄")}</td></tr>'
    mine = [p for p in pairs if p["sire"].startswith(i) or str(p["dam"]).startswith(i)]
    pr = ""
    if mine:
        pr = f'<section><div class="eyebrow">{bi("Pairing","配對")}</div><table>' + "".join(
            f'<tr><th>{esc(p["code"])}</th><td>{esc(p["sire"])} × {esc(p["dam"])}<br><span class="notes">{esc(p["goal"])} — {esc(p["status"])}</span></td></tr>'
            for p in mine) + "</table></section>"
    photo = (f'<div class="photo"><img src="{PHOTO_DIR}/{i.lower()}.jpg" alt="{esc(nm(a["name"]))}" '
             f'onerror="this.parentElement.style.display=\'none\'"></div>')
    title = f"{i} · {nm(a['name'])} — UMBRA LEACHIES"
    return (head(title, "records.html", f"{nm(a['name'])} — {a['lin']}", f"{PHOTO_DIR}/{i.lower()}.jpg", icon)
            + f'<header><div class="id">{i}</div><h1>{esc(nm(a["name"]))} <span class="sx">{sym(a["sex"])}</span></h1>'
            + f'<div class="sub">{esc(a["lin"])}</div><div class="chips">{chips}</div>'
            + (f'<div class="chips">{claims}</div>' if claims else "") + "</header>"
            + f'<section>{photo}<div class="eyebrow">{bi("Identity","身份")}</div><table>{tbl}</table></section>'
            + (f'<section><div class="eyebrow">{bi("Weight","體重")}</div>{cards}{spark(pts)}<table>{wl}</table></section>' if pts else "")
            + f'<section><div class="eyebrow">{bi("Feeding","餵食")}</div><table>{feed}</table></section>' + pr
            + f'<section><div class="eyebrow">{bi("Record notes","紀錄備註")}</div><p class="notes">{esc(a["notes"])}</p></section>'
            + foot())

def records_page(animals, icon):
    rows = ""
    for i, a in animals.items():
        chips = "".join(f'<span class="chip {"v" if s=="verified" else "u"}" title="{esc(en)}"></span>' for s, en, zh in CLAIMS.get(i, []))
        rows += (f'<tr><th><a href="./{i.lower()}.html">{i}</a></th><td><a href="./{i.lower()}.html">{esc(nm(a["name"]))} {sym(a["sex"])}</a>'
                 f'<br><span class="notes">{esc(a["lin"])}</span></td><td class="n">{a["wt"] if a["wt"] else "—"} g</td></tr>')
    n_l = sum(1 for k in animals if k.startswith("LCH")); n_h = sum(1 for k in animals if k.startswith("HOG"))
    return (head("Records — UMBRA LEACHIES", "records.html", "Collection records", None, icon)
            + f'<header><div class="id">{bi("COLLECTION","收藏")}</div><h1>{bi("Animal records","動物紀錄")}</h1>'
            + f'<div class="sub">{n_l} Rhacodactylus leachianus · {n_h} Heterodon nasicus</div></header>'
            + f'<section><p class="muted">{bi("Every animal here has a page with its lineage, weight history and feeding plan. Provenance claims are marked ✓ where documented and ? where they rest on a seller\'s word only.","每隻動物都有獨立頁面，載有血統、體重紀錄同餵食計劃。來源聲稱有文件支持嘅標 ✓，只係賣家口頭講嘅標 ?。")}</p></section>'
            + f'<section><table>{rows}</table></section>' + foot())

def available_page(animals, pairs, icon):
    items = AVAILABLE_OVERRIDE or [
        (p["code"], p["sire"], p["dam"], (str(p["dam_ready"]) if p["dam_ready"] else "—"), p["goal"], p["goal"])
        for p in pairs if str(p["status"]).lower() != "hold"]
    cards = "".join(
        f'<div class="card"><div class="k">{esc(code)} · {esc(exp)}</div><div class="v" style="font-size:20px">{esc(sire)} × {esc(dam)}</div>'
        f'<div class="t">{bi(esc(en),esc(zh))}</div></div>' for code, sire, dam, exp, en, zh in items)
    return (head("Available — UMBRA LEACHIES", "available.html", "Planned pairings and waitlist", None, icon)
            + f'<header><div class="id">{bi("AVAILABLE","預訂")}</div><h1>{bi("Nothing for sale yet.","暫時未有出售。")}</h1>'
            + f'<div class="sub">{bi("Planned pairings · waitlist open","計劃配對 · 接受排隊")}</div></header>'
            + f'<section><p>{bi("The collection is being grown to breeding weight. These are the pairings planned, with the earliest season each could produce. Offspring will be listed here first, with the same records their parents carry.","收藏正養至繁殖體重。以下係計劃中嘅配對，同各自最早可能出貨嘅季度。子代會先喺呢度上架，附帶同父母一樣完整嘅紀錄。")}</p></section>'
            + f'<section><div class="eyebrow">{bi("Planned pairings","計劃配對")}</div><div class="grid2">{cards}</div></section>'
            + f'<section><div class="eyebrow">{bi("Waitlist","排隊")}</div><p>{bi("Say which pairing you are interested in and you will hear first when eggs are down.","講明對邊個配對有興趣，落蛋時會第一時間通知你。")}</p>{contact_block()}</section>'
            + foot())

def home_page(animals, icon_big, icon):
    top = sorted(((k, v) for k, v in animals.items() if k.startswith("LCH") and v["val"]), key=lambda kv: -kv[1]["val"])[:4]
    feat = "".join(
        f'<a class="card" href="./{i.lower()}.html"><div class="k">{i}</div><div class="v" style="font-size:20px">{esc(nm(a["name"]))} {sym(a["sex"])}</div>'
        f'<div class="t">{esc(a["lin"])}</div></a>' for i, a in top)
    return (head("UMBRA LEACHIES · 玄影 — Rhacodactylus leachianus, Hong Kong", "index.html",
                 "A small Hong Kong collection of Rhacodactylus leachianus built around a documented melanistic line.", None, icon)
            + f'<div class="hero">{icon_big}<div class="w">UMBRA</div><div class="w2">LEACHIES</div><div class="zhw">玄影</div>'
            + f'<div class="tag">Rhacodactylus leachianus · Hong Kong · {bi("Melanistic project","黑化計劃")}</div></div>'
            + f'<section><div class="eyebrow">{bi("The project","計劃")}</div>'
            + f'<h2>{bi("Black, kept on the record.","黑，有紀錄可查。")}</h2>'
            + f'<p>{bi("Melanism in leachianus is a simple recessive: an animal shows it, carries it, or has neither. The line here starts from a registered visual male out of Ridiculous Rhacs. Around it sits a small collection of pure-locality animals — Caanawa, Nuu Ana, Nuu Ami, Brosse — chosen so that every clutch has a documented answer to the question <em>what is it</em>.","巨人守宮嘅黑化係簡單隱性基因：一隻動物要麼顯性、要麼帶因、要麼兩者都無。呢條線由一隻 Ridiculous Rhacs 出品、有登記嘅顯性雄開始。周圍係一批細而純 locality 嘅動物 —— Caanawa、Nuu Ana、Nuu Ami、Brosse —— 揀佢哋係為咗每一窩蛋對「呢隻係咩」呢個問題都有文件答到。")}</p>'
            + f'<p class="muted">{bi("Weights are logged monthly. Every animal has a page. Claims that rest on a seller\'s word are marked as such.","體重每月記錄，每隻動物有獨立頁面。只靠賣家口講嘅聲稱會明確標示。")}</p></section>'
            + f'<section><div class="eyebrow">{bi("Selected animals","精選")}</div><div class="grid2">{feat}</div>'
            + f'<p style="margin-top:18px"><a class="cta" href="./records.html">{bi("All records","全部紀錄")}</a><a class="cta" href="./available.html">{bi("Planned pairings","計劃配對")}</a></p></section>'
            + f'<section><div class="eyebrow">{bi("Contact","聯絡")}</div>{contact_block()}</section>'
            + foot())

# ------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=WORKBOOK); ap.add_argument("--out", default="site")
    args = ap.parse_args()
    if not os.path.exists(args.xlsx): sys.exit(f"workbook not found: {args.xlsx}")
    animals, weights, pairs = load(args.xlsx)
    os.makedirs(args.out, exist_ok=True)
    icon_raw = open(ICON_FILE, encoding="utf-8").read() if os.path.exists(ICON_FILE) else ""
    icon = re.sub(r'\s(width|height)="\d+"', "", icon_raw, count=2)
    for i, a in animals.items():
        open(os.path.join(args.out, f"{i.lower()}.html"), "w", encoding="utf-8").write(animal_page(i, a, weights, pairs, icon))
    open(os.path.join(args.out, "records.html"), "w", encoding="utf-8").write(records_page(animals, icon))
    open(os.path.join(args.out, "available.html"), "w", encoding="utf-8").write(available_page(animals, pairs, icon))
    open(os.path.join(args.out, "index.html"), "w", encoding="utf-8").write(home_page(animals, icon, icon))
    if icon_raw: open(os.path.join(args.out, ICON_FILE), "w", encoding="utf-8").write(icon_raw)
    print(f"wrote {len(animals)+3} pages to {args.out}/")

if __name__ == "__main__":
    main()
